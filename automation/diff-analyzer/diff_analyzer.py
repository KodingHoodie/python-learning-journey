import openpyxl

from openpyxl.styles import Font, numbers

wb = openpyxl.load_workbook('Employee Sales.xlsx')

while True:
    ws1_name = input('Please enter the name of the first worksheet: ')
    rng1_addr = input('Please specify the target range of cells on the first worksheet: ')

    ws2_name = input('Please enter the name of the second worksheet: ')
    rng2_addr = input('Please specify the target range of cells on the second worksheet: ')

    ws1 = wb[ws1_name]
    ws2 = wb[ws2_name]

    rng1 = ws1[rng1_addr]
    rng2 = ws2[rng2_addr]

    num_rows = len(rng1)

    num_cols = len(rng1[0])

    if num_rows == len(rng2) and num_cols == len(rng2[0]):
        break
    else:
        print('Please select two ranges with identical dimensions.')

headers = rng1[0]

diffs = []

for i in range(1, num_rows):
    for j in range(num_cols):
        cell1 = rng1[i][j].value
        cell2 = rng2[i][j].value

        if cell1 != cell2:
            diff = [i, headers[j].value, cell1, cell2]

            diffs.append(diff)

if 'Diffs' in wb.sheetnames:
    del wb['Diffs']

diffs_ws = wb.create_sheet('Diffs')

diffs_rng = diffs_ws['A2:D' + str(len(diffs) + 1)]

for i, row in enumerate(diffs_rng):
    for j, cell in enumerate(row):
        cell.value = diffs[i][j]

header_row = diffs_ws['A1:D1']

diffs_headers = ['Row', 'Column', 'Value 1', 'Value 2']

for idx, cell in enumerate(header_row[0]):
    cell.value = diffs_headers[idx]


formula_rng = diffs_ws['E2:E' + str(len(diffs) + 1)]

for idx, row in enumerate(formula_rng):
    current_row = str(idx + 2)

    diff_data = diffs[idx]

    if type(diff_data[2]) in(int, float) and type(diff_data[3]) in(int, float):
        cell1_addr = 'C' + current_row
        cell2_addr = 'D' + current_row

        formula = '=ABS(' + cell1_addr + '-' + cell2_addr + ')/AVERAGE(' + cell1_addr + ':' + cell2_addr + ')'

        target_cell = row[0]

        target_cell.value = formula

        target_cell.font = Font(color='FF0000')

        target_cell.number_format = numbers.FORMAT_PERCENTAGE



wb.save('Employee Sales.xlsx')